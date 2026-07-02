import os
import sys
import time
import re
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill

# Load environment variables from .env file
load_dotenv()

# Setup Configurations
USERNAME = os.getenv("SCRAPER_USERNAME")
PASSWORD = os.getenv("SCRAPER_PASSWORD")
LOGIN_URL = os.getenv("LOGIN_URL", "https://www.gleim.com/account/")
DATA_URL = os.getenv("DATA_URL", "https://www.gleim.com/testprep/ea/2019/part2?action=study")

# Selectors
SELECTOR_USERNAME = os.getenv("SELECTOR_USERNAME_INPUT", "#email")
SELECTOR_PASSWORD = os.getenv("SELECTOR_PASSWORD_INPUT", "#password")
SELECTOR_LOGIN_BTN = os.getenv("SELECTOR_LOGIN_BUTTON", "button.accountBtn")

# Excel Config
EXCEL_FILE = os.getenv("EXCEL_TEMPLATE_NAME", "gleim_questions.xlsx")
SHEET_NAME = os.getenv("EXCEL_SHEET_NAME", "Sheet1")

def get_element_text_with_tables(page, selector):
    """
    Retrieves the inner text of an element, converting any HTML <table>
    elements inside it into Markdown tables.
    """
    element = page.query_selector(selector)
    if not element:
        return ""
    
    text = page.evaluate("""
        (element) => {
            if (!element) return "";
            const clone = element.cloneNode(true);
            const tables = clone.querySelectorAll("table");
            tables.forEach(table => {
                let markdown = "\\n\\n";
                const rows = Array.from(table.querySelectorAll("tr"));
                let columnCount = 0;
                rows.forEach((row, rowIndex) => {
                    const cells = Array.from(row.querySelectorAll("th, td"));
                    if (rowIndex === 0) {
                        columnCount = cells.length;
                    }
                    const cellTexts = cells.map(cell => {
                        let text = cell.innerText.trim().replace(/\\n/g, " ");
                        if (cell.querySelector("strong, b")) {
                            text = `**${text}**`;
                        }
                        return text;
                    });
                    markdown += "| " + cellTexts.join(" | ") + " |\\n";
                    if (rowIndex === 0) {
                        const separators = Array(columnCount).fill("---");
                        markdown += "| " + separators.join(" | ") + " |\\n";
                    }
                });
                markdown += "\\n";
                const textNode = document.createTextNode(markdown);
                table.parentNode.replaceChild(textNode, table);
            });
            return clone.innerText.trim();
        }
    """, element)
    return text

def wait_for_explanation_to_contain(page, target_letter):
    """
    Waits for the answer explanation box to contain feedback for the target option letter.
    Converts tables inside the explanation to Markdown tables.
    """
    timeout = 5.0
    start_time = time.time()
    selector = ".answer-explanation .explanation-text"
    # Fallback to the main container if .explanation-text is not present
    if not page.query_selector(selector):
        selector = ".answer-explanation"
        
    while time.time() - start_time < timeout:
        try:
            exp_text = get_element_text_with_tables(page, selector)
            if f"answer ({target_letter.lower()}) is" in exp_text.lower():
                return exp_text
        except Exception:
            pass
        time.sleep(0.1)
    # Fallback to current text if timeout is reached
    try:
        return get_element_text_with_tables(page, selector)
    except Exception:
        return "Could not extract explanation"

def get_subunit_header(page):
    """
    Finds the subunit header text on the page using regex.
    """
    try:
        body_text = page.inner_text("body")
        for line in body_text.split('\n'):
            if "Study Unit" in line and "Subunit" in line:
                return line.strip()
    except Exception as e:
        print(f"Error getting subunit header: {e}")
    return ""

def login_and_scrape():
    # Basic validation
    if not USERNAME or not PASSWORD:
        print("Error: SCRAPER_USERNAME and SCRAPER_PASSWORD must be set in your .env file.")
        sys.exit(1)

    print("Starting Playwright...")
    with sync_playwright() as p:
        # Launch browser in non-headless mode so the user can see it and handle CAPTCHAs/MFA if needed
        print("Launching Chromium browser...")
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()

        # Step 1: Navigate to login page
        print(f"Navigating to login page: {LOGIN_URL}")
        page.goto(LOGIN_URL)
        page.wait_for_load_state("load")

        # Step 2: Log in
        print("Attempting automatic login...")
        try:
            page.fill(SELECTOR_USERNAME, USERNAME)
            page.fill(SELECTOR_PASSWORD, PASSWORD)
            print("Clicking login button...")
            page.click(SELECTOR_LOGIN_BTN)
            page.wait_for_load_state("load")
            time.sleep(2)
        except Exception as e:
            print("\n[Notice] Auto-login did not complete automatically.")
            print("👉 Please manually log in or complete authentication (CAPTCHA/MFA) in the browser window.")
            input("Once you are logged in and looking at your dashboard, press Enter here in the terminal to continue...")

        # Step 3: Navigate to study URL
        print(f"Navigating to study page: {DATA_URL}")
        try:
            page.goto(DATA_URL)
            page.wait_for_load_state("load")
            time.sleep(2)
        except Exception as e:
            print("Could not load study URL automatically. Please navigate to it manually in the browser.")
            input("Once you are on the Study page, press Enter here to continue...")

        # Let the user set up and start the study session manually
        print("\n--- Scraper ready ---")
        print("Please use the opened browser window to log in and start your study session.")
        print("Make sure you navigate to Question 1 of the study session before starting.")
        input("Once you see Question 1 on your screen, press Enter here to start scraping...")

        # Locate the page/tab that has the study session
        print("Scanning open tabs to find the active study session...")
        study_page = None
        for p_index, p_obj in enumerate(context.pages):
            try:
                p_obj.wait_for_load_state("load")
                print(f"Checking Tab {p_index}: Title='{p_obj.title()}', URL='{p_obj.url}'")
                if p_obj.query_selector("#queStem"):
                    study_page = p_obj
                    print(f"-> Selected Tab {p_index} (found #queStem)")
                    break
            except Exception:
                pass
        
        if not study_page:
            # Fallback to scanning URLs
            for p_index, p_obj in enumerate(context.pages):
                if "action=test" in p_obj.url or "study" in p_obj.url:
                    study_page = p_obj
                    print(f"-> Selected Tab {p_index} based on URL: {p_obj.url}")
                    break
        
        if not study_page:
            print("Warning: Could not automatically detect the tab containing the study session.")
            print("Defaulting to the last opened tab.")
            study_page = context.pages[-1]
        
        page = study_page
        scraped_questions = []
        question_counter = 1

        # Step 5: Scraping Loop
        while True:
            print(f"\n--- Scraping Question {question_counter} ---")
            
            # Wait for question stem to be visible
            try:
                page.wait_for_selector("#queStem", timeout=15000)
            except Exception:
                print("No question stem found (#queStem). Saving scraped data and stopping...")
                break

            # 1. Extract Question Stem
            question_text = get_element_text_with_tables(page, "#queStem")
            
            # 2. Extract Subunit Header
            subunit_header = get_subunit_header(page)
            print(f"Subunit: {subunit_header}")
            
            # Try to grab the question ID or number from active sidebar element if possible
            question_id = str(question_counter)
            try:
                # Look for sidebar items with marker or class
                # Standard Gleim sidebar lists questions
                active_item = page.query_selector(".activeQuestion, .sidebar-list a.selected, .questionsList a.active")
                if active_item:
                    question_id = active_item.inner_text().strip()
            except Exception:
                pass

            # 3. Process options
            option_letters = ['a', 'b', 'c', 'd']
            options_data = {}
            correct_answer = ""

            for letter in option_letters:
                click_selector = f'label[for="answer-{letter}"]'
                text_selector = f'#answer-{letter}-foil'
                
                # Check if option exists
                if not page.query_selector(click_selector):
                    continue
                
                # Extract option text
                option_text = ""
                if page.query_selector(text_selector):
                    option_text = get_element_text_with_tables(page, text_selector)
                
                # Click option to get explanation
                try:
                    # Scroll element into view if needed
                    page.locator(click_selector).scroll_into_view_if_needed()
                    page.click(click_selector)
                    
                    # Wait for explanation box to update to this option
                    exp_text = wait_for_explanation_to_contain(page, letter)
                    
                    is_correct = False
                    if "is correct" in exp_text.lower() or "anscorrect" in exp_text.lower():
                        is_correct = True
                        correct_answer = letter.upper()
                    
                    options_data[letter] = {
                        'text': option_text,
                        'explanation': exp_text,
                        'is_correct': is_correct
                    }
                    print(f"  Option {letter.upper()}: Extracted explanation. (Correct: {is_correct})")
                except Exception as e:
                    print(f"  Error extracting Option {letter.upper()}: {e}")
                    options_data[letter] = {
                        'text': option_text,
                        'explanation': "Error extracting explanation",
                        'is_correct': False
                    }
                
                time.sleep(0.3)

            # Store the scraped question data
            scraped_questions.append({
                'subunit': subunit_header,
                'question_id': question_id,
                'question_text': question_text,
                'options': options_data,
                'correct_answer': correct_answer
            })

            # Check if there is an active Next button
            next_btn = page.query_selector(".next-button")
            if next_btn and next_btn.is_visible() and next_btn.is_enabled():
                print("Clicking Next button...")
                next_btn.click()
                page.wait_for_load_state("load")
                question_counter += 1
                time.sleep(1) # Polite delay
                
                # Double check if the question text actually changed
                try:
                    new_question_text = get_element_text_with_tables(page, "#queStem")
                    if new_question_text == question_text:
                        print("Question text did not change after clicking Next. Finished scraping!")
                        break
                except Exception:
                    print("Could not find question stem on next page. Finished scraping!")
                    break
            else:
                print("No active Next button found. Scraping finished!")
                break

        print(f"\nScraped {len(scraped_questions)} total questions.")
        
        # Step 6: Save data to Excel
        if scraped_questions:
            # Determine Excel output name based on the first scraped question's subunit
            output_file = EXCEL_FILE
            first_subunit = scraped_questions[0]['subunit']
            if "Study Unit 18" in first_subunit:
                output_file = "study_unit_18.xlsx"
            elif "Study Unit 19" in first_subunit:
                output_file = "study_unit_19.xlsx"
            elif "Study Unit 20" in first_subunit:
                output_file = "study_unit_20.xlsx"
            else:
                match = re.search(r'Study Unit (\d+)', first_subunit)
                if match:
                    output_file = f"study_unit_{match.group(1)}.xlsx"
            
            save_questions_to_excel(scraped_questions, output_file)
        else:
            print("No questions scraped. Excel file was not modified.")

        print("Closing browser...")
        browser.close()

def save_questions_to_excel(questions, filename):
    """
    Saves the scraped questions and explanations into Excel.
    """
    if not os.path.exists(filename):
        if os.path.exists(EXCEL_FILE):
            print(f"Excel file '{filename}' not found. Copying template '{EXCEL_FILE}'...")
            wb = openpyxl.load_workbook(EXCEL_FILE)
            if SHEET_NAME in wb.sheetnames:
                ws = wb[SHEET_NAME]
            else:
                ws = wb.active
            # Clear existing data rows so we start clean for the new unit, but keep the headers and style
            if ws.max_row > 1:
                print(f"Clearing {ws.max_row - 1} existing data rows from template...")
                ws.delete_rows(2, ws.max_row - 1)
        else:
            print(f"Excel file '{filename}' and template '{EXCEL_FILE}' not found. Creating a new Excel workbook.")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
    else:
        print(f"Loading existing Excel file: {filename}")
        wb = openpyxl.load_workbook(filename)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            print(f"Sheet '{SHEET_NAME}' not found in the workbook. Using the active sheet.")
            ws = wb.active

    # Read existing questions to avoid duplicates
    existing_keys = set()
    if ws.max_row > 1:
        for r in range(2, ws.max_row + 1):
            sub = ws.cell(row=r, column=1).value
            qid = ws.cell(row=r, column=2).value
            if sub is not None and qid is not None:
                existing_keys.add((str(sub).strip(), str(qid).strip()))

    # Write Headers if empty
    headers = [
        "Subunit", "Question ID", "Question Text",
        "Option A", "Option A Explanation",
        "Option B", "Option B Explanation",
        "Option C", "Option C Explanation",
        "Option D", "Option D Explanation",
        "Correct Answer"
    ]
    
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        # Write headers directly to row 1
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=1, column=col_idx + 1, value=header)
            # Apply premium styling to headers
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Classic Navy
            cell.font = header_font
            cell.fill = header_fill

    current_row = ws.max_row + 1
    skipped_count = 0
    added_count = 0

    # Write each question's row
    for q in questions:
        key = (str(q['subunit']).strip(), str(q['question_id']).strip())
        if key in existing_keys:
            skipped_count += 1
            continue
        
        row_values = [
            q['subunit'],
            q['question_id'],
            q['question_text'],
            q['options'].get('a', {}).get('text', ''),
            q['options'].get('a', {}).get('explanation', ''),
            q['options'].get('b', {}).get('text', ''),
            q['options'].get('b', {}).get('explanation', ''),
            q['options'].get('c', {}).get('text', ''),
            q['options'].get('c', {}).get('explanation', ''),
            q['options'].get('d', {}).get('text', ''),
            q['options'].get('d', {}).get('explanation', ''),
            q['correct_answer']
        ]
        for col_idx, value in enumerate(row_values):
            ws.cell(row=current_row, column=col_idx + 1, value=value)
        
        existing_keys.add(key)
        current_row += 1
        added_count += 1

    print(f"Excel saving summary: Added {added_count} new questions, skipped {skipped_count} duplicates.")
    wb.save(filename)
    print(f"Successfully saved data to '{filename}'!")

if __name__ == "__main__":
    login_and_scrape()
